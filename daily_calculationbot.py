import os
import re
import sys
import json
import time
import datetime
import logging
import requests
from flask import Flask
import threading
import io
from dotenv import load_dotenv

app = Flask(__name__)

@app.route('/')
def home():
    return "I'm alive!"

def run():
    app.run(host='0.0.0.0', port=int(os.environ.get("PORT", 8080)))

# Run the web server in a separate thread so it doesn't block the bot
threading.Thread(target=run).start()

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

load_dotenv()

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

TELEGRAM_BOT_TOKEN = os.getenv('TELEGRAM_BOT_TOKEN')
TELEGRAM_CHAT_ID   = os.getenv('TELEGRAM_CHAT_ID')

DATA_DIR = 'user_data'
os.makedirs(DATA_DIR, exist_ok=True)

# ── NLP ────────────────────────────────────────────────────────────────────
EXPENSE_KEYWORDS = [
    'bought','spent','paid','ate','drank','purchased','got','ordered',
    'subscribed','rented','hired','fee','bill','cost','price',
    'sotib','xarajat','yedim','ichdim',"to'ladim",'berdim',
]
INCOME_KEYWORDS = [
    'earned','received','got paid','salary','income','profit','bonus',
    'dividend','refund','sold','transfer','deposited','freelance',
    'oldim','ishlab topdim','maosh','daromad','sotdim','topdim',
]
TODO_KEYWORDS = [
    'should','need to','have to','must','going to','want to',
    'plan to','will',"i'll",'remember to',"don't forget",
    'kerak','borish kerak','qilish kerak','eslatma','unutma',
    'boray','qilay','boraman','qilaman','eslab qol','yodda tut',
]

AMOUNT_PATTERN = re.compile(
    r"(\d[\d\s,\.]*\d|\d+)\s*(sum|so\'m|som|uzs)?", re.IGNORECASE
)

def parse_amount(text):
    for raw, _ in AMOUNT_PATTERN.findall(text):
        cleaned = raw.replace(' ','').replace(',','').replace('.','')
        try:    return float(cleaned)
        except: continue
    return None

def parse_natural_language(text):
    tl     = text.lower()
    amount = parse_amount(tl)
    if amount is None:
        return None, None, text
    is_expense = any(kw in tl for kw in EXPENSE_KEYWORDS)
    is_income  = any(kw in tl for kw in INCOME_KEYWORDS)
    tx_type = 'income' if (is_income and not is_expense) else 'expense'
    return tx_type, amount, text

def is_todo_message(text):
    return any(kw in text.lower() for kw in TODO_KEYWORDS)

# ── Per-User Data ───────────────────────────────────────────────────────────
def _user_file(cid):
    return os.path.join(DATA_DIR, f'{cid}.json')

def load_user_data(cid):
    path = _user_file(cid)
    if os.path.exists(path):
        try:
            with open(path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except json.JSONDecodeError:
            logger.warning(f"Fayl buzilgan ({cid}), yangidan boshlanmoqda.")
    return {}

def save_user_data(cid, data):
    with open(_user_file(cid), 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

def get_today():
    return datetime.date.today().strftime('%Y-%m-%d')

def ensure_today(data):
    today = get_today()
    if today not in data:
        data[today] = {
            'total_income':  0,
            'total_expense': 0,
            'balance':       0,
            'transactions':  [],
            'todos':         [],
        }
    if 'todos' not in data[today]:
        data[today]['todos'] = []
    return today

# ── Transactions ────────────────────────────────────────────────────────────
def add_transaction(tx_type, amount, description, data):
    today = ensure_today(data)
    if tx_type == 'income':
        data[today]['total_income'] += amount
        data[today]['balance']      += amount
    else:
        data[today]['total_expense'] += amount
        data[today]['balance']       -= amount
    data[today]['transactions'].append({
        'type':        tx_type,
        'amount':      amount,
        'description': description,
        'timestamp':   datetime.datetime.now().isoformat(),
    })
    balance = data[today]['balance']
    sign    = '+' if balance >= 0 else ''
    if tx_type == 'income':
        mood = '\U0001f929 Barakalla!' if amount >= 500000 else '\U0001f60a Ajoyib!'
        return (
            f"\U0001f4b0 *Daromad qo'shildi!* {mood}\n"
            f"\U0001f4dd _{description}_\n"
            f"\U0001f4b5 Miqdor: *{amount:,.0f} so'm*\n\n"
            f"\u2696\ufe0f Bugungi balans: *{sign}{balance:,.0f} so'm* "
            + ('\U0001f60d' if balance >= 0 else '\U0001f622')
        )
    else:
        mood = "\U0001f622 Ko'p ketdi!" if amount >= 200000 else '\U0001f605 Yaxshi!'
        return (
            f"\U0001f4b8 *Xarajat qo'shildi!* {mood}\n"
            f"\U0001f4dd _{description}_\n"
            f"\U0001f4b5 Miqdor: *{amount:,.0f} so'm*\n\n"
            f"\u2696\ufe0f Bugungi balans: *{sign}{balance:,.0f} so'm* "
            + ('\U0001f60a' if balance >= 0 else '\U0001f630')
        )

# ── To-Do ───────────────────────────────────────────────────────────────────
def add_todo(task_text, data):
    today = ensure_today(data)
    todos = data[today]['todos']
    todos.append({'task': task_text, 'done': False,
                  'created': datetime.datetime.now().isoformat()})
    idx = len(todos)
    return (
        f"\u2705 *Vazifa qo'shildi!* \U0001f4aa\n"
        f"_{task_text}_\n\n"
        f"\U0001f4cb Siz bugun *{idx} ta* vazifa qo'shdingiz. Zo'r!"
    )

def list_todos(data):
    today = ensure_today(data)
    todos = data[today]['todos']
    if not todos:
        return (
            "\U0001f634 Bugun hali hech qanday vazifa yo'q!\n\n"
            "\U0001f4a1 Kichik qadam \u2014 katta natija! Birinchi vazifangizni qo'sing \U0001f31f"
        )
    done_count = sum(1 for t in todos if t['done'])
    total      = len(todos)
    lines      = [f"\U0001f4aa *Bugungi Vazifalar* ({done_count}/{total} bajarildi)\n"]
    for i, t in enumerate(todos, 1):
        check = '\u2705' if t['done'] else '\u2b1c'
        lines.append(f"{check} {i}. {t['task']}")
    if done_count == total and total > 0:
        lines.append("\n\U0001f389 *Barakalla! Barcha vazifalar bajarildi!* \U0001f3c6")
    elif done_count > 0:
        lines.append(f"\n\U0001f525 *{done_count} ta bajarildi, davom eting!*")
    else:
        lines.append("\n\U0001f4a1 _Birinchi vazifani boshlash vaqti!_")
    return '\n'.join(lines)

def complete_todo(num, data):
    today = ensure_today(data)
    todos = data[today]['todos']
    if num < 1 or num > len(todos):
        return f"\U0001f615 #{num} vazifa topilmadi. Sizda {len(todos)} ta vazifa bor."
    todos[num - 1]['done'] = True
    return (
        f"\U0001f389 *Barakalla!* #{num} vazifa bajarildi! \U0001f4aa\n\n"
        f"\u2705 _{todos[num-1]['task']}_\n\n"
        "Davom eting, siz zo'rsiz! \U0001f525"
    )

def delete_todo(num, data):
    today = ensure_today(data)
    todos = data[today]['todos']
    if num < 1 or num > len(todos):
        return f"\U0001f615 #{num} vazifa topilmadi."
    removed = todos.pop(num - 1)
    return f"\U0001f5d1\ufe0f #{num} vazifa o'chirildi.\n_'{removed['task']}'_\n\n\U0001f60c Yaxshi qaror!"

# ── Summary ─────────────────────────────────────────────────────────────────
def show_summary(data):
    today   = ensure_today(data)
    d       = data[today]
    balance = d['balance']
    sign    = '+' if balance >= 0 else ''
    bal_emoji = '\U0001f60d' if balance > 100000 else ('\U0001f60a' if balance >= 0 else '\U0001f630')
    lines = [
        f"\U0001f4ca *Bugungi Hisobot* \U0001f4c5 {today}\n",
        f"\U0001f4b0 Daromad:  *{d['total_income']:,.0f} so'm* \U0001f929",
        f"\U0001f4b8 Xarajat:  *{d['total_expense']:,.0f} so'm* \U0001f624",
        f"\u2696\ufe0f Balans:   *{sign}{balance:,.0f} so'm* {bal_emoji}",
    ]
    if d['transactions']:
        lines.append("\n\U0001f4cb *Bugungi harakatlar:*")
        for tx in d['transactions'][-10:]:
            tt   = tx.get('type','expense')
            desc = tx.get('description','')
            if tt == 'note':
                lines.append(f"\U0001f4dd _(eslatma)_ {desc}")
            else:
                icon = '\U0001f7e2' if tt == 'income' else '\U0001f534'
                lines.append(f"{icon} {tx['amount']:,.0f} so'm \u2014 _{desc}_")
    if d['todos']:
        done  = sum(1 for t in d['todos'] if t['done'])
        total = len(d['todos'])
        lines.append(f"\n\u2705 *Vazifalar:* {done}/{total} bajarildi "
                     + ('\U0001f3c6' if done == total else '\U0001f4aa'))
        for i, t in enumerate(d['todos'], 1):
            check = '\u2705' if t['done'] else '\u2b1c'
            lines.append(f"{check} {i}. {t['task']}")
    if not d['transactions'] and not d['todos']:
        lines.append("\n\U0001f305 _Bugun hali hech narsa yo'q. Yangi kun \u2014 yangi imkoniyat!_ \u2728")
    return '\n'.join(lines)

def show_weekly_summary(data):
    today    = datetime.date.today()
    start    = today - datetime.timedelta(days=today.weekday())
    total_in = total_ex = 0
    lines    = ["\U0001f4ca *Haftalik Hisobot* \U0001f4c5\n"]
    for i in range(7):
        d = (start + datetime.timedelta(days=i)).strftime('%Y-%m-%d')
        if d in data:
            day_data = data[d]
            total_in += day_data['total_income']
            total_ex += day_data['total_expense']
            bal  = day_data['balance']
            sign = '+' if bal >= 0 else ''
            emoji = '\U0001f60a' if bal >= 0 else '\U0001f630'
            lines.append(f"\U0001f4c5 {d}: {sign}{bal:,.0f} so'm {emoji}")
    bal  = total_in - total_ex
    sign = '+' if bal >= 0 else ''
    lines += [
        f"\n\U0001f4b0 Jami Daromad:   *{total_in:,.0f} so'm* \U0001f929",
        f"\U0001f4b8 Jami Xarajat:   *{total_ex:,.0f} so'm* \U0001f624",
        f"\u2696\ufe0f Haftalik Balans: *{sign}{bal:,.0f} so'm* "
        + ('\U0001f60d' if bal >= 0 else '\U0001f630'),
    ]
    return '\n'.join(lines)

def show_monthly_summary(data):
    today    = datetime.date.today()
    month    = today.strftime('%Y-%m')
    total_in = total_ex = 0
    for d, di in data.items():
        if d.startswith(month):
            total_in += di['total_income']
            total_ex += di['total_expense']
    bal  = total_in - total_ex
    sign = '+' if bal >= 0 else ''
    return (
        f"\U0001f4ca *Oylik Hisobot* \U0001f5d3\ufe0f {month}\n\n"
        f"\U0001f4b0 Daromad:  *{total_in:,.0f} so'm* \U0001f929\n"
        f"\U0001f4b8 Xarajat:  *{total_ex:,.0f} so'm* \U0001f624\n"
        f"\u2696\ufe0f Balans:   *{sign}{bal:,.0f} so'm* "
        + ('\U0001f60d' if bal >= 0 else '\U0001f630')
    )

# ── Excel ────────────────────────────────────────────────────────────────────
def generate_excel(data, period='today'):
    if not EXCEL_AVAILABLE:
        return None, "openpyxl o'rnatilmagan. Bajaring: pip install openpyxl"

    wb = openpyxl.Workbook()

    header_font  = Font(bold=True, color='FFFFFF', size=11)
    income_fill  = PatternFill('solid', fgColor='27AE60')
    expense_fill = PatternFill('solid', fgColor='E74C3C')
    note_fill    = PatternFill('solid', fgColor='3498DB')
    header_fill  = PatternFill('solid', fgColor='2C3E50')
    alt_fill     = PatternFill('solid', fgColor='ECF0F1')
    center       = Alignment(horizontal='center', vertical='center')
    thin         = Side(style='thin', color='BDC3C7')
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    def style_header(cell, fill=None):
        cell.font      = header_font
        cell.fill      = fill or header_fill
        cell.alignment = center
        cell.border    = border

    def style_cell(cell, alt=False):
        if alt: cell.fill = alt_fill
        cell.alignment = center
        cell.border    = border

    today = get_today()
    if period == 'today':
        dates = [today]; title = f"Kunlik Hisobot \u2014 {today}"
    elif period == 'week':
        base  = datetime.date.today()
        start = base - datetime.timedelta(days=base.weekday())
        dates = [(start + datetime.timedelta(days=i)).strftime('%Y-%m-%d') for i in range(7)]
        title = f"Haftalik Hisobot \u2014 {start}"
    elif period == 'month':
        base  = datetime.date.today(); month = base.strftime('%Y-%m')
        dates = sorted([d for d in data if d.startswith(month)])
        title = f"Oylik Hisobot \u2014 {month}"
    else:
        dates = sorted(data.keys()); title = "Barcha Hisobot"

    ws = wb.active
    ws.title = "Tranzaksiyalar"; ws.sheet_properties.tabColor = "2C3E50"
    ws.merge_cells('A1:F1')
    ws['A1'].value     = f"\U0001f4b0 {title}"
    ws['A1'].font      = Font(bold=True, size=14, color='FFFFFF')
    ws['A1'].fill      = header_fill; ws['A1'].alignment = center
    ws.row_dimensions[1].height = 30

    headers = ["Sana","Vaqt","Tur","Miqdor (so'm)","Tavsif","Balans"]
    ws.append([]); ws.row_dimensions[2].height = 5
    ws.append(headers)
    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=3, column=col, value=h))
    ws.row_dimensions[3].height = 22

    row_num = 4; running_balance = 0; txs = []
    for date in dates:
        if date in data:
            for tx in data[date]['transactions']:
                ts = tx.get('timestamp','')
                try:   tstr = datetime.datetime.fromisoformat(ts).strftime('%H:%M')
                except: tstr = ''
                txs.append((date, tstr, tx))

    for i, (date, tstr, tx) in enumerate(txs):
        alt = (i % 2 == 1); tt = tx.get('type','expense')
        if tt == 'income':   running_balance += tx['amount']; lbl = "\U0001f4e5 Daromad"
        elif tt == 'note':   lbl = "\U0001f4dd Eslatma"
        else:                running_balance -= tx['amount']; lbl = "\U0001f4e4 Xarajat"
        ws.append([date, tstr, lbl, tx['amount'] if tt != 'note' else 0,
                   tx.get('description',''), running_balance])
        for col in range(1, 7): style_cell(ws.cell(row=row_num, column=col), alt)
        tc      = ws.cell(row=row_num, column=3)
        tc.fill = income_fill if tt == 'income' else (note_fill if tt == 'note' else expense_fill)
        tc.font = Font(bold=True, color='FFFFFF')
        ws.cell(row=row_num, column=4).number_format = '#,##0'
        ws.cell(row=row_num, column=6).number_format = '#,##0'
        row_num += 1

    for col, w in zip('ABCDEF', [14,10,14,18,40,18]):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("Hisobot"); ws2.sheet_properties.tabColor = "27AE60"
    ws2.merge_cells('A1:C1')
    ws2['A1'].value = "\U0001f4ca Kunlar bo'yicha Hisobot"
    ws2['A1'].font  = Font(bold=True, size=13, color='FFFFFF')
    ws2['A1'].fill  = header_fill; ws2['A1'].alignment = center
    ws2.row_dimensions[1].height = 28
    ws2.append([]); ws2.append(["Sana","Daromad (so'm)","Xarajat (so'm)","Balans (so'm)"])
    for col in range(1, 5): style_header(ws2.cell(row=3, column=col))
    ws2.row_dimensions[3].height = 22

    total_in = total_ex = 0
    for i, date in enumerate(dates):
        if date in data:
            day_data = data[date]; total_in += day_data['total_income']; total_ex += day_data['total_expense']
            ws2.append([date, day_data['total_income'], day_data['total_expense'], day_data['balance']])
            ri = 4 + i
            for col in range(1,5): style_cell(ws2.cell(row=ri, column=col), i % 2 == 1)
            for col in [2,3,4]: ws2.cell(row=ri, column=col).number_format = '#,##0'
            ws2.cell(row=ri, column=4).font = Font(bold=True,
                color='27AE60' if day_data['balance'] >= 0 else 'E74C3C')

    tr = 4 + len(dates)
    ws2.append(['JAMI', total_in, total_ex, total_in - total_ex])
    for col in range(1,5):
        c = ws2.cell(row=tr, column=col)
        c.font = Font(bold=True, color='FFFFFF'); c.fill = header_fill
        c.border = border; c.alignment = center
    for col in [2,3,4]: ws2.cell(row=tr, column=col).number_format = '#,##0'
    for col, w in zip('ABCD', [14,18,18,18]): ws2.column_dimensions[col].width = w

    ws3 = wb.create_sheet("Vazifalar"); ws3.sheet_properties.tabColor = "E67E22"
    ws3.merge_cells('A1:C1')
    ws3['A1'].value = "\u2705 Vazifalar Ro'yxati"
    ws3['A1'].font  = Font(bold=True, size=13, color='FFFFFF')
    ws3['A1'].fill  = PatternFill('solid', fgColor='E67E22'); ws3['A1'].alignment = center
    ws3.row_dimensions[1].height = 28
    ws3.append([]); ws3.append(['#','Vazifa','Holat','Sana'])
    for col in range(1,5):
        style_header(ws3.cell(row=3,column=col), PatternFill('solid',fgColor='E67E22'))
    ws3.row_dimensions[3].height = 22

    task_row = 4
    for date in dates:
        if date in data:
            for j, todo in enumerate(data[date].get('todos',[]), 1):
                status = '\u2705 Bajarildi' if todo['done'] else '\u2b1c Kutilmoqda'
                ws3.append([j, todo['task'], status, date])
                for col in range(1,5): style_cell(ws3.cell(row=task_row,column=col), task_row % 2 == 1)
                ws3.cell(row=task_row,column=3).font = Font(
                    color='27AE60' if todo['done'] else 'E74C3C', bold=True)
                task_row += 1
    for col, w in zip('ABCD', [6,45,18,14]): ws3.column_dimensions[col].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return buf, None

# ── Telegram Bot ─────────────────────────────────────────────────────────────
class TelegramBot:
    def __init__(self):
        self.user_cache     = {}
        self.user_states    = {}
        self.last_update_id = 0
        self.base_url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}"
        logger.info("Bot ishga tushdi \u2705")

    STATE_IDLE            = 'idle'
    STATE_TODO_ADD        = 'todo_add'
    STATE_JOURNAL_INCOME  = 'journal_income'
    STATE_JOURNAL_EXPENSE = 'journal_expense'
    STATE_JOURNAL_NOTE    = 'journal_note'

    def get_state(self, c): return self.user_states.get(c, self.STATE_IDLE)
    def set_state(self, c, s): self.user_states[c] = s
    def clear_state(self, c): self.user_states[c] = self.STATE_IDLE

    def get_data(self, cid):
        if cid not in self.user_cache:
            self.user_cache[cid] = load_user_data(cid)
        return self.user_cache[cid]

    def save(self, cid): save_user_data(cid, self.user_cache[cid])

    # ── Keyboards ─────────────────────────────────────────────────────────
    def kb_main(self):
        return {"inline_keyboard": [
            [{"text":"\U0001f4cb Vazifalar",      "callback_data":"todo_menu"},
             {"text":"\U0001f4d3 Kundalik",        "callback_data":"journal_menu"}],
            [{"text":"\U0001f4ca Hisobot",         "callback_data":"cmd_summary"},
             {"text":"\U0001f4e4 Excel Yuklash",   "callback_data":"cmd_excel"}],
        ]}

    def kb_todo(self):
        return {"inline_keyboard": [
            [{"text":"\u2795 Vazifa Qo'shish",        "callback_data":"todo_add"}],
            [{"text":"\U0001f4cb Vazifalarni Ko'rish", "callback_data":"todo_view"}],
            [{"text":"\u270f\ufe0f O'zgartirish",      "callback_data":"todo_changes"}],
            [{"text":"\U0001f3e0 Asosiy Menyu",        "callback_data":"main_menu"}],
        ]}

    def kb_journal(self):
        return {"inline_keyboard": [
            [{"text":"\U0001f4b0 Daromad Qo'shish (+)",           "callback_data":"journal_income"}],
            [{"text":"\U0001f4b8 Xarajat Qo'shish (\u2212)",      "callback_data":"journal_expense"}],
            [{"text":"\U0001f4dd Eslatma Yozish (o'zgarishsiz)",  "callback_data":"journal_note"}],
            [{"text":"\U0001f3e0 Asosiy Menyu",                    "callback_data":"main_menu"}],
        ]}

    def kb_changes(self, todos):
        rows = []
        for i, t in enumerate(todos, 1):
            icon  = '\u2705' if t['done'] else '\u2b1c'
            label = t['task'][:22] + '\u2026' if len(t['task']) > 22 else t['task']
            rows.append([
                {"text":f"{icon} {i}. {label}", "callback_data":f"todo_toggle_{i}"},
                {"text":"\U0001f5d1\ufe0f O'chir",  "callback_data":f"todo_del_{i}"},
            ])
        rows.append([{"text":"\U0001f519 Orqaga",      "callback_data":"todo_menu"},
                     {"text":"\U0001f3e0 Asosiy Menyu","callback_data":"main_menu"}])
        return {"inline_keyboard": rows}

    def kb_back(self):
        return {"inline_keyboard": [[{"text":"\U0001f3e0 Asosiy Menyu","callback_data":"main_menu"}]]}

    def kb_after_todo(self):
        return {"inline_keyboard": [
            [{"text":"\u2795 Yana Qo'shish",          "callback_data":"todo_add"}],
            [{"text":"\U0001f4cb Vazifalarni Ko'rish", "callback_data":"todo_view"}],
            [{"text":"\U0001f3e0 Asosiy Menyu",        "callback_data":"main_menu"}],
        ]}

    def kb_after_tx(self):
        return {"inline_keyboard": [
            [{"text":"\U0001f4b0 Daromad","callback_data":"journal_income"},
             {"text":"\U0001f4b8 Xarajat","callback_data":"journal_expense"}],
            [{"text":"\U0001f4d3 Kundalik","callback_data":"journal_menu"},
             {"text":"\U0001f3e0 Asosiy",  "callback_data":"main_menu"}],
        ]}

    # ── Telegram API ──────────────────────────────────────────────────────
    def delete_webhook(self):
        try:
            r = requests.post(f"{self.base_url}/deleteWebhook",
                              data={'drop_pending_updates': False}, timeout=10)
            if r.json().get('result'): logger.info("Webhook o'chirildi \u2705")
        except Exception as e: logger.warning(f"Webhook xato: {e}")

    def get_updates(self):
        try:
            r = requests.get(f"{self.base_url}/getUpdates",
                             params={'offset': self.last_update_id + 1, 'timeout': 30},
                             timeout=35)
            r.raise_for_status(); return r.json().get('result', [])
        except requests.exceptions.HTTPError as e:
            if e.response is not None and e.response.status_code == 409:
                logger.warning("409 Conflict \u2014 webhook o'chirilmoqda...")
                self.delete_webhook(); time.sleep(5)
            else: logger.error(f"get_updates HTTP xato: {e}")
            return []
        except Exception as e: logger.error(f"get_updates xato: {e}"); return []

    def send_msg(self, cid, text, kb=None):
        try:
            d = {'chat_id': cid, 'text': text, 'parse_mode': 'Markdown'}
            if kb: d['reply_markup'] = json.dumps(kb)
            requests.post(f"{self.base_url}/sendMessage", data=d, timeout=10)
        except Exception as e: logger.error(f"send_msg xato: {e}")

    def edit_msg(self, cid, mid, text, kb=None):
        try:
            d = {'chat_id': cid, 'message_id': mid, 'text': text, 'parse_mode': 'Markdown'}
            if kb: d['reply_markup'] = json.dumps(kb)
            requests.post(f"{self.base_url}/editMessageText", data=d, timeout=10)
        except Exception as e: logger.error(f"edit_msg xato: {e}")

    def answer_cb(self, cbid):
        try:
            requests.post(f"{self.base_url}/answerCallbackQuery",
                          data={'callback_query_id': cbid}, timeout=10)
        except Exception as e: logger.error(f"answer_cb xato: {e}")

    def send_doc(self, cid, buf, fname, caption=''):
        try:
            requests.post(f"{self.base_url}/sendDocument",
                          data={'chat_id': cid, 'caption': caption, 'parse_mode': 'Markdown'},
                          files={'document': (fname, buf,
                              'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')},
                          timeout=30)
        except Exception as e: logger.error(f"send_doc xato: {e}")

    # ── Callback handler ──────────────────────────────────────────────────
    def handle_cb(self, update):
        cq = update.get('callback_query')
        if not cq: return
        cbid = cq['id']; cid = cq['message']['chat']['id']
        mid  = cq['message']['message_id']; cb = cq.get('data','')
        self.answer_cb(cbid); self.last_update_id = update['update_id']
        data = self.get_data(cid); ensure_today(data)

        if cb == 'main_menu':
            self.clear_state(cid)
            self.edit_msg(cid, mid,
                "\U0001f3e0 *Asosiy Menyu* \U0001f60a\n\nNima qilmoqchisiz?", self.kb_main())

        elif cb == 'todo_menu':
            self.clear_state(cid)
            self.edit_msg(cid, mid,
                "\U0001f4cb *Vazifalar Ro'yxati* \U0001f4aa\n\nNima qilmoqchisiz?", self.kb_todo())

        elif cb == 'todo_add':
            self.set_state(cid, self.STATE_TODO_ADD)
            self.edit_msg(cid, mid,
                "\U0001f4dd *Vazifa Qo'shish* \u270d\ufe0f\n\n"
                "Vazifangizni *yozing* \U0001f447\n\n"
                "_Misol: Oziq-ovqat sotib olish_ \U0001f6d2\n"
                "_Misol: Soat 18:00 da onaga qo'ng'iroq_ \U0001f4de")

        elif cb == 'todo_view':
            bk = {"inline_keyboard": [
                [{"text":"\u270f\ufe0f O'zgartirish","callback_data":"todo_changes"}],
                [{"text":"\U0001f519 Orqaga",      "callback_data":"todo_menu"},
                 {"text":"\U0001f3e0 Asosiy Menyu","callback_data":"main_menu"}],
            ]}
            self.edit_msg(cid, mid, list_todos(data), bk)

        elif cb == 'todo_changes':
            todos = data[get_today()]['todos']
            if not todos:
                self.edit_msg(cid, mid,
                    "\U0001f634 Hali vazifalar yo'q!\n\nBirinchi vazifangizni qo'sing \U0001f31f",
                    {"inline_keyboard": [
                        [{"text":"\u2795 Vazifa Qo'shish","callback_data":"todo_add"}],
                        [{"text":"\U0001f3e0 Asosiy Menyu","callback_data":"main_menu"}],
                    ]})
            else:
                self.edit_msg(cid, mid,
                    "\u270f\ufe0f *O'zgartirish* \U0001f527\n\n"
                    "Vazifani bosib *\u2705/\u2b1c o'zgartiring*\n"
                    "*\U0001f5d1\ufe0f O'chir* tugmasini bosib o'chiring:",
                    self.kb_changes(todos))

        elif cb.startswith('todo_toggle_'):
            num = int(cb.split('_')[2]); today = get_today()
            todos = data[today]['todos']
            if 1 <= num <= len(todos):
                todos[num-1]['done'] = not todos[num-1]['done']; self.save(cid)
            self.edit_msg(cid, mid,
                "\u270f\ufe0f *O'zgartirish* \U0001f527\n\n"
                "Vazifani bosib *\u2705/\u2b1c o'zgartiring*\n"
                "*\U0001f5d1\ufe0f O'chir* tugmasini bosib o'chiring:",
                self.kb_changes(data[today]['todos']))

        elif cb.startswith('todo_del_'):
            num = int(cb.split('_')[2]); today = get_today()
            todos = data[today]['todos']
            if 1 <= num <= len(todos): todos.pop(num-1); self.save(cid)
            remaining = data[today]['todos']
            if remaining:
                self.edit_msg(cid, mid,
                    "\u270f\ufe0f *O'zgartirish* \U0001f527\n\n"
                    "Vazifani bosib *\u2705/\u2b1c o'zgartiring*\n"
                    "*\U0001f5d1\ufe0f O'chir* tugmasini bosib o'chiring:",
                    self.kb_changes(remaining))
            else:
                self.edit_msg(cid, mid,
                    "\U0001f5d1\ufe0f Barcha vazifalar o'chirildi! \U0001f60c\n\n"
                    "Yangi vazifa qo'shmoqchimisiz? \U0001f4aa",
                    {"inline_keyboard": [
                        [{"text":"\u2795 Vazifa Qo'shish","callback_data":"todo_add"}],
                        [{"text":"\U0001f3e0 Asosiy Menyu","callback_data":"main_menu"}],
                    ]})

        elif cb == 'journal_menu':
            self.clear_state(cid)
            self.edit_msg(cid, mid,
                "\U0001f4d3 *Kundalik* \U0001f4b0\n\nNima qilmoqchisiz?", self.kb_journal())

        elif cb == 'journal_income':
            self.set_state(cid, self.STATE_JOURNAL_INCOME)
            self.edit_msg(cid, mid,
                "\U0001f4b0 *Daromad Qo'shish* \U0001f929\n\n"
                "*Miqdor* va tavsifni yozing:\n\n"
                "_Misollar:_\n"
                "\u2022 `500000 maosh` \U0001f4bc\n"
                "\u2022 `200000 freelance` \U0001f4bb\n"
                "\u2022 `1000000 bonus` \U0001f381")

        elif cb == 'journal_expense':
            self.set_state(cid, self.STATE_JOURNAL_EXPENSE)
            self.edit_msg(cid, mid,
                "\U0001f4b8 *Xarajat Qo'shish* \U0001f605\n\n"
                "*Miqdor* va tavsifni yozing:\n\n"
                "_Misollar:_\n"
                "\u2022 `50000 taksi` \U0001f695")

        elif cb == 'journal_note':
            self.set_state(cid, self.STATE_JOURNAL_NOTE)
            self.edit_msg(cid, mid,
                "\U0001f4dd *Eslatma Yozish* \u2728\n\n"
                "Istalgan narsani yozing \u2014\n"
                "*balansga hech narsa ta'sir qilmaydi* \U0001f60a\n\n"
                "_Misollar:_\n"
                "\u2022 _20 so'mga kola ichdim_ \U0001f964\n"
                "\u2022 _Bugungi tushlik juda mazali edi_ \U0001f60b\n"
                "\u2022 _Do'stga 5000 berdim_ \U0001f91d")

        elif cb == 'cmd_summary':
            self.edit_msg(cid, mid, show_summary(data), self.kb_back())

        elif cb == 'cmd_excel':
            self.edit_msg(cid, mid,
                "\U0001f4e4 *Excel Yuklash* \U0001f4ca\n\nDavrni tanlang:",
                {"inline_keyboard": [
                    [{"text":"\U0001f4c5 Bugun",       "callback_data":"excel_today"},
                     {"text":"\U0001f4c6 Bu Hafta",    "callback_data":"excel_week"}],
                    [{"text":"\U0001f5d3\ufe0f Bu Oy", "callback_data":"excel_month"},
                     {"text":"\U0001f4da Hammasi",     "callback_data":"excel_all"}],
                    [{"text":"\U0001f3e0 Asosiy Menyu","callback_data":"main_menu"}],
                ]})

        elif cb.startswith('excel_'):
            period = cb.split('_',1)[1]
            names  = {'today':'Bugun','week':'Bu Hafta','month':'Bu Oy','all':'Hammasi'}
            self.edit_msg(cid, mid,
                f"\u23f3 *{names.get(period,period)}* uchun Excel tayyorlanmoqda\u2026 \U0001f4ca",
                self.kb_back())
            self.do_excel(cid, data, period)

    # ── Message handler ───────────────────────────────────────────────────
    def handle_message(self, update):
        if 'callback_query' in update:
            self.handle_cb(update); return
        try:
            msg  = update.get('message') or update.get('edited_message')
            if not msg: return
            cid  = msg['chat']['id']
            text = msg.get('text','').strip()
            self.last_update_id = update['update_id']
            if not text: return
            data  = self.get_data(cid)
            state = self.get_state(cid)

            if state == self.STATE_TODO_ADD:
                self.clear_state(cid); ensure_today(data)
                self.send_msg(cid, add_todo(text, data), self.kb_after_todo())
                self.save(cid); return

            elif state == self.STATE_JOURNAL_INCOME:
                self.clear_state(cid); ensure_today(data)
                amount = parse_amount(text)
                if not amount:
                    self.send_msg(cid,
                        "\U0001f615 *Miqdor topilmadi!*\n\n"
                        "_Raqam kiriting. Misol: `500000 maosh`_ \U0001f4b5",
                        {"inline_keyboard": [
                            [{"text":"\U0001f501 Qayta Urinish","callback_data":"journal_income"},
                             {"text":"\U0001f3e0 Asosiy Menyu", "callback_data":"main_menu"}],
                        ]})
                else:
                    self.send_msg(cid, add_transaction('income', amount, text, data),
                                  self.kb_after_tx())
                    self.save(cid)
                return

            elif state == self.STATE_JOURNAL_EXPENSE:
                self.clear_state(cid); ensure_today(data)
                amount = parse_amount(text)
                if not amount:
                    self.send_msg(cid,
                        "\U0001f615 *Miqdor topilmadi!*\n\n"
                        "_Raqam kiriting. Misol: `20000 hot-dog`_ \U0001f32d",
                        {"inline_keyboard": [
                            [{"text":"\U0001f501 Qayta Urinish","callback_data":"journal_expense"},
                             {"text":"\U0001f3e0 Asosiy Menyu", "callback_data":"main_menu"}],
                        ]})
                else:
                    self.send_msg(cid, add_transaction('expense', amount, text, data),
                                  self.kb_after_tx())
                    self.save(cid)
                return

            elif state == self.STATE_JOURNAL_NOTE:
                self.clear_state(cid); today = ensure_today(data)
                data[today]['transactions'].append({
                    'type': 'note', 'amount': 0,
                    'description': text,
                    'timestamp': datetime.datetime.now().isoformat(),
                })
                self.save(cid)
                self.send_msg(cid,
                    f"\U0001f4dd *Eslatma saqlandi!* \u2728\n\n_{text}_\n\n"
                    "\U0001f49a _(Balans o'zgarmadi \u2014 bu shunchaki yodgorlik)_",
                    {"inline_keyboard": [
                        [{"text":"\U0001f4dd Yana Eslatma","callback_data":"journal_note"}],
                        [{"text":"\U0001f4d3 Kundalik",    "callback_data":"journal_menu"},
                         {"text":"\U0001f3e0 Asosiy",      "callback_data":"main_menu"}],
                    ]})
                return

            resp = self.route(text, cid, data)
            if resp:
                if isinstance(resp, tuple): self.send_msg(cid, resp[0], resp[1])
                else: self.send_msg(cid, resp)
        except Exception as e:
            logger.error(f"handle_message xato: {e}")

    # ── Command router ────────────────────────────────────────────────────
    def route(self, text, cid, data):
        cmd = text.split()[0].lower()
        if   cmd in ('/start','/menu'):      return self.cmd_start(cid)
        elif cmd == '/help':                  return self.cmd_help()
        elif cmd in ('/summary','/hisobot'): return show_summary(data)
        elif cmd in ('/week','/hafta'):       return show_weekly_summary(data)
        elif cmd in ('/month','/oy'):         return show_monthly_summary(data)
        elif cmd in ('/excel','/export'):
            parts  = text.split()
            period = parts[1].lower() if len(parts) > 1 else 'today'
            if period not in ('today','week','month','all'): period = 'today'
            self.do_excel(cid, data, period); return None
        elif cmd in ('/todo','/vazifa'):
            task = text.split(' ',1)[1].strip() if ' ' in text else ''
            if not task:
                return ("\U0001f60a Vazifani kiriting:\n"
                        "`/vazifa Oziq-ovqat sotib olish`\n\n"
                        "Yoki /menu ni oching")
            ensure_today(data); r = add_todo(task, data); self.save(cid); return r
        elif cmd == '/todos': return list_todos(data)
        elif cmd == '/done':
            parts = text.split()
            if len(parts) < 2 or not parts[1].isdigit():
                return "\U0001f60a Foydalanish: `/done <raqam>`\nMisol: `/done 1`"
            ensure_today(data); r = complete_todo(int(parts[1]), data)
            self.save(cid); return r
        elif cmd == '/deltodo':
            parts = text.split()
            if len(parts) < 2 or not parts[1].isdigit():
                return "\U0001f60a Foydalanish: `/deltodo <raqam>`"
            ensure_today(data); r = delete_todo(int(parts[1]), data)
            self.save(cid); return r
        elif cmd in ('/income','/daromad'):
            rest   = text.split(' ',1)[1].strip() if ' ' in text else ''
            amount = parse_amount(rest)
            if not amount: return "\U0001f60a Foydalanish: `/daromad 500000 maosh`"
            ensure_today(data)
            r = add_transaction('income', amount, rest or 'Daromad', data)
            self.save(cid); return r
        elif cmd in ('/expense','/xarajat'):
            rest   = text.split(' ',1)[1].strip() if ' ' in text else ''
            amount = parse_amount(rest)
            if not amount: return "\U0001f60a Foydalanish: `/xarajat 20000 hot-dog`"
            ensure_today(data)
            r = add_transaction('expense', amount, rest or 'Xarajat', data)
            self.save(cid); return r
        else: return self.nlp(text, cid, data)

    def nlp(self, text, cid, data):
        tx_type, amount, desc = parse_natural_language(text)
        if amount is None:
            if is_todo_message(text):
                ensure_today(data); r = add_todo(text, data)
                self.save(cid); return r
            return (
                "\U0001f914 *Hmm, tushunmadim...*\n\n"
                "\n\n"
                "\U0001f4b8  \U0001f32d\n"
                "\U0001f4b0  \U0001f4bc\n"
                "\U0001f4cb  \U0001f393\n\n"
                "Yoki menyuni oching: /menu \U0001f60a"
            )
        ensure_today(data); r = add_transaction(tx_type, amount, desc, data)
        self.save(cid); return r

    # ── Excel sender ──────────────────────────────────────────────────────
    def do_excel(self, cid, data, period):
        if not EXCEL_AVAILABLE:
            self.send_msg(cid,
                "\u274c openpyxl o'rnatilmagan.\nBajaring: `pip install openpyxl`")
            return
        buf, err = generate_excel(data, period)
        if err: self.send_msg(cid, f"\u274c {err}"); return
        today  = get_today()
        names  = {'today':'bugun','week':'hafta','month':'oy','all':'hammasi'}
        fname  = f"hisobot_{names.get(period,period)}_{today}.xlsx"
        caption = f"\U0001f4ca *Excel Hisobot* \U0001f389 \u2014 {period}\n_{today}_"
        self.send_doc(cid, buf, fname, caption)
        logger.info(f"Excel yuborildi: {fname}")

    # ── /start ────────────────────────────────────────────────────────────
    def cmd_start(self, cid=None):
        if cid is not None:
            data  = self.get_data(cid)
            today = ensure_today(data)
            bal   = data[today]['balance']
            sign  = '+' if bal >= 0 else ''
            bal_e = '\U0001f60d' if bal >= 0 else '\U0001f630'
            return (
                f"\U0001f389 *Kunlik Moliya Botiga xush kelibsiz!* \U0001f31f\n\n"
                f"\u2696\ufe0f Bugungi balansiz: *{sign}{bal:,.0f} so'm* {bal_e}\n\n"
                "Quyidagi tugmalardan birini bosing \U0001f447",
                self.kb_main()
            )
        return "Boshlash uchun /start bosing. \U0001f60a"

    def cmd_help(self):
        return (
            "\U0001f4d6 *Barcha Buyruqlar* \U0001f60a\n\n"
            "\u2501\u2501\u2501\u2501 \U0001f3e0 *Menyu* \u2501\u2501\u2501\u2501\n"
            "  `/menu` yoki `/start` \u2014 Asosiy menyuni ochish\n\n"
            "\u2501\u2501\u2501\u2501 \U0001f4b0 *Pul* \u2501\u2501\u2501\u2501\n"
            "Tabiiy tilda:\n"
            "  _hot-dog uchun 20000 to'ladim_ \U0001f32d\n"
            "  _maosh 5000000 oldim_ \U0001f4bc\n\n"
            "Yoki buyruqlar:\n"
            "  ``\n"
            "  ``\n\n"
            "\u2501\u2501\u2501\u2501 \U0001f4cb *Vazifalar* \u2501\u2501\u2501\u2501\n"
            "  `/vazifa <matn>` \u2014 Qo'shish \u270d\ufe0f\n"
            "  `/todos` \u2014 Ko'rish \U0001f440\n"
            "  `/done <n>` \u2014 Bajarildi \u2705\n"
            "  `/deltodo <n>` \u2014 O'chirish \U0001f5d1\ufe0f\n\n"
            "\u2501\u2501\u2501\u2501 \U0001f4ca *Hisobotlar* \u2501\u2501\u2501\u2501\n"
            "  `/hisobot` | `/hafta` | `/oy`\n"
            "  `/excel` \u2014 Excel (bugun) \U0001f4e5\n"
            "  `/excel week` / `/excel all`\n"
        )

    # ── Run loop ──────────────────────────────────────────────────────────
    def run(self):
        logger.info("Webhook o'chirilmoqda...")
        self.delete_webhook(); time.sleep(1)
        logger.info("Bot ishlamoqda... To'xtatish uchun Ctrl+C bosing.")
        self.send_msg(
            TELEGRAM_CHAT_ID,
            "\U0001f680 *Bot ishga tushdi!* \U0001f389\n\n"
            "Salom! Men sizning moliyaviy yordamchingizman! \U0001f60a\n\n"
            "Menyuni ochish uchun /menu bosing \U0001f447",
            self.kb_main()
        )
        while True:
            try:
                updates = self.get_updates()
                for upd in updates: self.handle_message(upd)
                time.sleep(0.5)
            except KeyboardInterrupt:
                logger.info("Bot to'xtatildi. Xayr! \U0001f44b"); break
            except Exception as e:
                logger.error(f"Bot loop xato: {e}"); time.sleep(5)

# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    if not TELEGRAM_BOT_TOKEN:
        logger.error("TELEGRAM_BOT_TOKEN .env faylida topilmadi!")
        sys.exit(1)
    TelegramBot().run()

if __name__ == '__main__':
    main()